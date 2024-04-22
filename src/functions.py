import cv2
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter


# マークシートからマークを検出
def get_marks(image, lower_color, upper_color) -> list[tuple[tuple[int, int], int]]:
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
    blur = cv2.GaussianBlur(hsv, (9, 9), 0)
    color = cv2.inRange(blur, lower_color, upper_color)

    element8 = np.array([[1, 1, 1], [1, 1, 1], [1, 1, 1]], np.uint8)
    oc = cv2.morphologyEx(color, cv2.MORPH_OPEN, element8)
    oc = cv2.morphologyEx(oc, cv2.MORPH_CLOSE, element8)

    contours, _ = cv2.findContours(oc, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    marks = []
    for contour in contours:
        (x, y), radius = cv2.minEnclosingCircle(contour)
        # 誤検出を除外
        if radius > 10:
            marks.append(((int(x), int(y)), int(radius)))
        
    print(f"{len(marks)}個のマークを検出しました")
    return marks


# マークの座標をもとに二次元配列を作成
def get_activities(filename, rows, columns) -> list[list[bool]]:
    cell_width, cell_height = 150, 50
    
    image = cv2.imread(filename)
    _, image = cv2.threshold(image, 128, 255, cv2.THRESH_BINARY)
    image = cv2.resize(image, (rows * cell_width, columns * cell_height))

    marks = get_marks(image, np.array([0, 0, 0]), np.array([96, 96, 96]))
    activities = [[False] * rows for _ in range(columns)]

    for mark in marks:
        x = mark[0][0] // cell_width
        y = mark[0][1] // cell_height
        activities[y][x] = True
        
    plt.imshow(image)
    plt.show()
    return activities


def write_to_namelist(workbook, activities, start_row, start_column) -> None:
    for row, line in enumerate(activities):
        for column, cell in enumerate(line):
            if cell:
                workbook["名簿"].cell(row=row + start_row, column=column + start_column, value=1)

                
def write_to_report(workbook, days, all_member_count, start_row, start_column) -> None:
    for day in range(start_column, start_column + days):
        letter = get_column_letter(day)
        cell_range = f"{letter}{start_row}:{letter}{start_row + all_member_count}"
        values = [cell.value for row in workbook["名簿"][cell_range] for cell in row if isinstance(cell.value, (int, float))]
        member_count = sum(values)
        
        if member_count != 0:
            workbook["報告書"].cell(row=day + 1, column=4, value=member_count)
            workbook["報告書"].cell(row=day + 1, column=5, value="音研部室")
            workbook["報告書"].cell(row=day + 1, column=6, value="バンド練習")